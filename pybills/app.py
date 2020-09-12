from openpyxl import Workbook, load_workbook
from datetime import date

wb_name = 'contas.xlsx'

try:
    wb = load_workbook(wb_name)
except FileNotFoundError:
    wb = Workbook()

description = input('Insira a descrição da movimentação: ')

valid = False
while not valid:
    try:
        value = float(input('Digite o valor (inteiro ou separado por "."): '))
        valid = True
    except ValueError:
        print('Erro: Valor deve ser numérico inteiro, ou separado por "."')
        valid = False

meses = (
    'Janeiro',
    'Fevereiro',
    'Março',
    'Abril',
    'Maio',
    'Junho',
    'Julho',
    'Agosto',
    'Setembro',
    'Outubro',
    'Novembro',
    'Dezembro'
)

today = date.today()
mes_atual = meses[int(today.strftime('%m')) - 1]
ano = today.strftime('%Y')
sheet_name = f'{mes_atual}-{ano}'

try:
    ws = wb[sheet_name]
except KeyError:
    ws = wb.create_sheet(title=sheet_name, index=(int(today.strftime('%m'))))
    ws['A1'] = 'Data'
    ws['B1'] = 'Descrição'
    ws['C1'] = 'Valor'

new_row = ws.max_row
if not ws[f'A{new_row}'].value == 'Total':
    new_row += 1

ws[f'A{new_row}'] = today.strftime('%d/%m/%Y')
ws[f'B{new_row}'] = description
ws[f'C{new_row}'] = value

total_spent = 0

for row in range(new_row):
    if row <= 1:
        continue
    total_spent += float(ws[f'C{row}'].value)
total_spent += value

total_row = new_row+1
ws[f'A{total_row}'] = 'Total'
ws[f'B{total_row}'] = total_spent

wb.save(filename=wb_name)
